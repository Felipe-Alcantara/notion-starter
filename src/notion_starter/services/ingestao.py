"""Ingestão extensível de fontes externas para o Notion.

Fontes apenas coletam e normalizam itens. O caso de uso :func:`ingerir`
coordena criação/atualização no Notion e mantém o fluxo idempotente pela
propriedade ``Origem``.
"""

from __future__ import annotations

import os
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Protocol, runtime_checkable

from notion_starter import NotionClient, properties

LIMITE_TEXTO_NOTION = 2000
EXTENSOES_TEXTO = {
    ".css",
    ".csv",
    ".html",
    ".ini",
    ".js",
    ".json",
    ".md",
    ".py",
    ".toml",
    ".ts",
    ".txt",
    ".xml",
    ".yaml",
    ".yml",
}


@runtime_checkable
class Fonte(Protocol):
    """Contrato implementado por qualquer origem de dados ingerível."""

    def coletar(self) -> Iterable[ItemColetado]:
        """Itera sobre os itens disponíveis na fonte."""
        ...


@dataclass
class ItemColetado:
    """Item normalizado produzido por uma fonte.

    ``propriedades`` carrega valores de propriedade **já no formato do Notion**
    (montados com :mod:`notion_starter.properties`); fontes tabulares como a
    :class:`FontePlanilha` usam este campo para que cada coluna vire uma
    propriedade tipada da linha, e não só texto.
    """

    nome: str
    tipo_fonte: str
    conteudo: str = ""
    metadados: dict[str, Any] = field(default_factory=dict)
    origem: str = ""
    propriedades: dict[str, Any] = field(default_factory=dict)


@dataclass
class ResultadoIngestao:
    """Resumo de uma operação de ingestão em lote.

    ``falhas`` descreve cada item que falhou (``"nome (origem): motivo"``),
    na ordem em que os erros aconteceram — sem isso, ``erros`` é um contador
    cego e o chamador não tem como diagnosticar o lote.
    """

    criados: int = 0
    atualizados: int = 0
    erros: int = 0
    itens_processados: int = 0
    falhas: list[str] = field(default_factory=list)


class FonteArquivos:
    """Coleta arquivos de uma pasta, com metadados e prévia textual segura."""

    def __init__(
        self,
        pasta: str | Path,
        *,
        extensoes: list[str] | None = None,
        recursivo: bool = True,
        max_caracteres: int = LIMITE_TEXTO_NOTION,
    ) -> None:
        if max_caracteres < 0:
            raise ValueError("max_caracteres não pode ser negativo.")
        self._pasta = Path(pasta)
        self._extensoes = (
            {self._normalizar_extensao(ext) for ext in extensoes} if extensoes is not None else None
        )
        self._recursivo = recursivo
        self._max_caracteres = max_caracteres

    @staticmethod
    def _normalizar_extensao(extensao: str) -> str:
        limpa = extensao.strip().lower()
        if not limpa:
            raise ValueError("extensoes não pode conter valores vazios.")
        return limpa if limpa.startswith(".") else f".{limpa}"

    def coletar(self) -> Iterable[ItemColetado]:
        if not self._pasta.is_dir():
            return

        candidatos = self._pasta.rglob("*") if self._recursivo else self._pasta.iterdir()
        for arquivo in sorted(candidatos):
            if arquivo.is_symlink() or not arquivo.is_file():
                continue
            extensao = arquivo.suffix.lower()
            if self._extensoes is not None and extensao not in self._extensoes:
                continue

            tamanho = arquivo.stat().st_size
            origem = arquivo.relative_to(self._pasta).as_posix()
            conteudo = self._conteudo(arquivo, tamanho)
            yield ItemColetado(
                nome=arquivo.name,
                tipo_fonte="arquivos",
                conteudo=conteudo,
                metadados={
                    "extensao": extensao,
                    "tamanho_bytes": tamanho,
                    "caminho_relativo": origem,
                },
                origem=origem,
            )

    def _conteudo(self, arquivo: Path, tamanho: int) -> str:
        resumo = f"Arquivo {arquivo.suffix or 'sem extensão'} ({tamanho} bytes)"
        if arquivo.suffix.lower() not in EXTENSOES_TEXTO or self._max_caracteres == 0:
            return resumo
        try:
            previa = arquivo.read_text(encoding="utf-8", errors="replace")[
                : self._max_caracteres
            ].strip()
        except OSError:
            return resumo
        return previa or resumo


class FonteGitHub:
    """Converte repositórios de um usuário em :class:`ItemColetado`."""

    def __init__(self, usuario: str, *, github_client: Any = None) -> None:
        self._usuario = usuario
        self._github_client = github_client

    def _resolver_client(self) -> Any:
        if self._github_client is not None:
            return self._github_client
        from notion_starter.github import GitHubClient

        return GitHubClient()

    def coletar(self) -> Iterable[ItemColetado]:
        for repo in self._resolver_client().listar_repos(self._usuario):
            yield ItemColetado(
                nome=repo.nome,
                tipo_fonte="github",
                conteudo=repo.descricao or "",
                metadados={
                    "linguagem": repo.linguagem,
                    "estrelas": repo.estrelas,
                    "forks": repo.forks,
                    "topicos": repo.topicos,
                    "privado": repo.privado,
                    "url_html": repo.url_html,
                    "atualizado_em": repo.atualizado_em,
                },
                origem=repo.nome_completo,
            )


#: Tipos de coluna aceitos pela :class:`FontePlanilha` e seus builders.
TIPOS_COLUNA = ("texto", "numero", "data", "select", "email", "url", "checkbox", "telefone")

#: Valores textuais tratados como "verdadeiro" numa coluna ``checkbox``.
_VALORES_VERDADEIROS = {"sim", "true", "verdadeiro", "x", "1", "yes", "feito", "ok"}


class FontePlanilha:
    """Converte uma planilha (``.xlsx``/``.csv``) em :class:`ItemColetado`.

    A primeira linha é o cabeçalho. Cada linha vira um item cujo nome sai da
    coluna-título (por padrão, a primeira) e cujas demais colunas viram
    propriedades tipadas via ``tipos`` (``coluna -> tipo``). Valores que não
    convertem para o tipo pedido (número/data inválidos) **não são
    descartados**: vão para a propriedade "Observações" da linha.

    Números e datas aceitam o formato brasileiro (``1.614``, ``2,7 mil``,
    ``dd/mm/aaaa``, serial do Excel) via :mod:`notion_starter.valores_br`.

    Args:
        caminho: Arquivo ``.xlsx`` (requer ``openpyxl``, extra ``planilha``)
            ou ``.csv`` (stdlib).
        aba: Nome da aba do ``.xlsx``; por padrão, a aba ativa.
        coluna_titulo: Coluna usada como título da linha; por padrão, a
            primeira do cabeçalho.
        tipos: Mapeamento ``coluna -> tipo`` (ver :data:`TIPOS_COLUNA`).
            Colunas fora do mapeamento viram ``rich_text``.
        renomear: Mapeamento opcional ``coluna -> nome da propriedade`` no
            Notion (por padrão, o próprio cabeçalho).
    """

    def __init__(
        self,
        caminho: str | Path,
        *,
        aba: str | None = None,
        coluna_titulo: str | None = None,
        tipos: dict[str, str] | None = None,
        renomear: dict[str, str] | None = None,
    ) -> None:
        self._caminho = Path(caminho)
        self._aba = aba
        self._coluna_titulo = coluna_titulo
        self._tipos = dict(tipos or {})
        self._renomear = dict(renomear or {})
        for coluna, tipo in self._tipos.items():
            if tipo not in TIPOS_COLUNA:
                raise ValueError(
                    f"Tipo '{tipo}' inválido para a coluna '{coluna}'. "
                    f"Tipos aceitos: {', '.join(TIPOS_COLUNA)}."
                )

    def coletar(self) -> Iterable[ItemColetado]:
        cabecalho, linhas = self._ler()
        if not cabecalho:
            return
        coluna_titulo = self._coluna_titulo or cabecalho[0]
        if coluna_titulo not in cabecalho:
            raise ValueError(f"Coluna-título '{coluna_titulo}' não existe no cabeçalho.")

        for numero, linha in enumerate(linhas, start=2):
            valores = dict(zip(cabecalho, linha, strict=False))
            titulo = str(valores.get(coluna_titulo) or "").strip()
            if not titulo:
                continue
            props, observacoes = self._propriedades_da_linha(valores, coluna_titulo)
            if observacoes:
                props["Observações"] = properties.rich_text(
                    _limitar_texto("; ".join(observacoes))
                )
            yield ItemColetado(
                nome=titulo,
                tipo_fonte="planilha",
                metadados={"linha": numero, "arquivo": self._caminho.name},
                origem=f"{self._caminho.name}:{numero}",
                propriedades=props,
            )

    # -- Leitura -------------------------------------------------------------

    def _ler(self) -> tuple[list[str], list[list[Any]]]:
        if not self._caminho.is_file():
            raise ValueError(f"Planilha não encontrada: {self._caminho}")
        sufixo = self._caminho.suffix.lower()
        if sufixo == ".csv":
            return self._ler_csv()
        if sufixo == ".xlsx":
            return self._ler_xlsx()
        raise ValueError(f"Formato não suportado: '{sufixo}'. Use .xlsx ou .csv.")

    def _ler_csv(self) -> tuple[list[str], list[list[Any]]]:
        import csv

        with self._caminho.open(encoding="utf-8-sig", newline="") as arquivo:
            leitor = csv.reader(arquivo)
            linhas = [linha for linha in leitor if any(str(c).strip() for c in linha)]
        if not linhas:
            return [], []
        cabecalho = [str(c).strip() for c in linhas[0]]
        return cabecalho, [list(linha) for linha in linhas[1:]]

    def _ler_xlsx(self) -> tuple[list[str], list[list[Any]]]:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - depende do ambiente
            raise ValueError(
                "Ler .xlsx requer o pacote openpyxl. Instale com: "
                "pip install 'notion-starter[planilha]'"
            ) from exc

        pasta = openpyxl.load_workbook(self._caminho, data_only=True, read_only=True)
        try:
            if self._aba is not None:
                if self._aba not in pasta.sheetnames:
                    raise ValueError(
                        f"Aba '{self._aba}' não existe. Abas: {', '.join(pasta.sheetnames)}."
                    )
                planilha = pasta[self._aba]
            else:
                planilha = pasta.active
            linhas = [
                list(linha)
                for linha in planilha.iter_rows(values_only=True)
                if any(c is not None and str(c).strip() for c in linha)
            ]
        finally:
            pasta.close()
        if not linhas:
            return [], []
        cabecalho = [str(c or "").strip() for c in linhas[0]]
        return cabecalho, linhas[1:]

    # -- Conversão -----------------------------------------------------------

    def _propriedades_da_linha(
        self, valores: dict[str, Any], coluna_titulo: str
    ) -> tuple[dict[str, Any], list[str]]:
        from notion_starter.valores_br import data_br, numero_br

        props: dict[str, Any] = {}
        observacoes: list[str] = []
        for coluna, bruto in valores.items():
            if coluna == coluna_titulo or not coluna:
                continue
            if bruto is None or not str(bruto).strip():
                continue
            texto = str(bruto).strip()
            nome_prop = self._renomear.get(coluna, coluna)
            tipo = self._tipos.get(coluna, "texto")

            if tipo == "numero":
                numero = numero_br(bruto)
                if numero is None:
                    observacoes.append(f"{coluna}: {texto}")
                else:
                    props[nome_prop] = properties.number(numero)
            elif tipo == "data":
                data = data_br(bruto)
                if data is None:
                    observacoes.append(f"{coluna}: {texto}")
                else:
                    props[nome_prop] = properties.date(data)
            elif tipo == "select":
                props[nome_prop] = properties.select(_limitar_texto(texto))
            elif tipo == "email":
                props[nome_prop] = properties.email(texto)
            elif tipo == "url":
                props[nome_prop] = properties.url(texto)
            elif tipo == "telefone":
                props[nome_prop] = properties.phone_number(texto)
            elif tipo == "checkbox":
                props[nome_prop] = properties.checkbox(texto.lower() in _VALORES_VERDADEIROS)
            else:
                props[nome_prop] = properties.rich_text(_limitar_texto(texto))
        return props, observacoes


def _limitar_texto(valor: str) -> str:
    return valor[:LIMITE_TEXTO_NOTION]


def _schema_do_database(client: NotionClient, database_id: str) -> dict[str, Any] | None:
    """Propriedades do database de destino, ou ``None`` se não der para ler.

    Sem o schema (clientes antigos, falha de rede), a ingestão mantém o
    comportamento clássico de enviar ``Nome``/``Fonte``/``Origem`` fixos.
    """
    try:
        schema = client.get_database(database_id).get("properties")
    except Exception:
        return None
    return schema or None


def _nome_da_propriedade_titulo(schema: dict[str, Any]) -> str:
    for nome, definicao in schema.items():
        if isinstance(definicao, dict) and definicao.get("type") == "title":
            return nome
    return "Nome"


def _propriedades_de_item(
    item: ItemColetado, schema: dict[str, Any] | None = None
) -> dict[str, Any]:
    nome = item.nome.strip()
    tipo_fonte = item.tipo_fonte.strip()
    if not nome:
        raise ValueError("ItemColetado.nome não pode estar vazio.")
    if not tipo_fonte:
        raise ValueError("ItemColetado.tipo_fonte não pode estar vazio.")

    # Com o schema em mãos, o título vai para a propriedade-título real do
    # database e os campos genéricos só entram se a coluna existir — sem isso,
    # um database sem "Nome"/"Fonte"/"Origem" rejeitava a linha inteira.
    nome_titulo = _nome_da_propriedade_titulo(schema) if schema else "Nome"
    props: dict[str, Any] = {nome_titulo: properties.title(_limitar_texto(nome))}
    if schema is None or "Fonte" in schema:
        props["Fonte"] = properties.select(tipo_fonte)
    if item.conteudo and (schema is None or "Descrição" in schema):
        props["Descrição"] = properties.rich_text(_limitar_texto(item.conteudo))
    if item.origem and (schema is None or "Origem" in schema):
        props["Origem"] = properties.rich_text(_limitar_texto(item.origem))
    # Propriedades tipadas da fonte por último: são mais específicas e podem
    # sobrescrever os campos genéricos (exceto os reservados acima).
    for nome_prop, valor in item.propriedades.items():
        if nome_prop not in (nome_titulo, "Fonte", "Origem"):
            props[nome_prop] = valor
    return props


def _pagina_por_origem(
    client: NotionClient,
    database_id: str,
    origem: str,
) -> dict[str, Any] | None:
    if not origem:
        return None
    paginas = client.consultar_database(
        database_id,
        page_size=1,
        filtro={
            "property": "Origem",
            "rich_text": {"equals": _limitar_texto(origem)},
        },
    )
    return paginas[0] if paginas else None


def ingerir(
    fonte: Fonte,
    *,
    client: NotionClient | None = None,
    database_id: str | None = None,
) -> ResultadoIngestao:
    """Cria ou atualiza no Notion os itens produzidos por ``fonte``."""

    if client is None:
        from integrations.notion import criar_cliente

        client = criar_cliente()

    db_id = database_id or os.environ.get("NOTION_DATABASE_ID", "").strip()
    if not db_id:
        raise ValueError(
            "database_id é obrigatório. Passe como argumento ou defina "
            "NOTION_DATABASE_ID no ambiente."
        )

    schema = _schema_do_database(client, db_id)
    # Sem a coluna "Origem" no destino não há chave de idempotência: filtrar
    # por ela devolveria 400, então cada item é criado direto.
    upsert_por_origem = schema is None or "Origem" in schema

    resultado = ResultadoIngestao()
    for item in fonte.coletar():
        resultado.itens_processados += 1
        try:
            props = _propriedades_de_item(item, schema)
            existente = (
                _pagina_por_origem(client, db_id, item.origem) if upsert_por_origem else None
            )
            if existente and existente.get("id"):
                client.atualizar_pagina(str(existente["id"]), props)
                resultado.atualizados += 1
            else:
                client.criar_pagina(db_id, props)
                resultado.criados += 1
        except Exception as exc:
            # Ingestão é lote: uma fonte inválida ou item rejeitado não impede
            # os itens seguintes. O resumo mantém a falha observável.
            resultado.erros += 1
            identificacao = item.nome.strip() or "(sem nome)"
            if item.origem:
                identificacao = f"{identificacao} ({item.origem})"
            # Exceções sem mensagem (ex.: ``KeyError`` vazio) ainda precisam
            # de um motivo legível no resumo.
            motivo = str(exc) or type(exc).__name__
            resultado.falhas.append(f"{identificacao}: {motivo}")

    return resultado
